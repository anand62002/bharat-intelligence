"""
data/fetchers.py — Bharat Intelligence India Market Data Fetchers
All functions return None on failure and log errors to stderr.
"""

import json
import logging
import os
import random
import re
import time
from datetime import datetime
from typing import Any, Callable, Optional, Tuple, Type

import feedparser
import requests
import yfinance as yf
from bs4 import BeautifulSoup

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

# Rotate User-Agent for screener.in requests to reduce compound UA-based blocking.
# Note: Railway's static IP is still the primary block vector — UA rotation alone
# won't bypass IP-level blocks, but prevents UA fingerprinting layering on top.
_SCREENER_USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) Gecko/20100101 Firefox/124.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
]


def _screener_headers(extra: dict | None = None) -> dict:
    """Return request headers with a randomised User-Agent for screener.in calls."""
    h = {
        **_HEADERS,
        "User-Agent": random.choice(_SCREENER_USER_AGENTS),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Encoding": "gzip, deflate",
        "Connection": "keep-alive",
    }
    if extra:
        h.update(extra)
    return h


# ── Screener.in session warmup ──────────────────────────────────────────────
# screener.in is Django-based and sets a csrftoken cookie + sessionid on first
# visit.  Making a direct /company/{slug}/ request without those cookies looks
# like a bot and triggers 403s from their Cloudflare/WAF layer.
# We keep a module-level session that is warmed once per process (homepage visit)
# and reused for all subsequent company-page requests.  If the session ages out
# it self-heals by re-warming on the next call.
_screener_session: requests.Session | None = None
_screener_session_warmed: bool = False
_screener_session_logged_in: bool = False   # True once login POST succeeded


def _screener_inject_session(session: "requests.Session") -> bool:
    """
    Inject a pre-existing screener.in session cookie into the requests session.

    screener.in supports Google OAuth login (and email/password login), both of
    which complete in a browser — there is no server-side credentials flow we can
    automate.  Instead, the user copies the ``sessionid`` cookie from their browser
    after logging in via Google and stores it in the ``SCREENER_SESSION`` env var.
    This is the same pattern used for Trendlyne (TRENDLYNE_SESSION).

    How to get the cookie:
      1. Log in at https://www.screener.in (via Google or email)
      2. Open DevTools → Application → Cookies → screener.in
      3. Copy the value of the ``sessionid`` cookie
      4. Set ``SCREENER_SESSION=<value>`` on Railway (and in your local .env)

    The cookie typically stays valid for several weeks.  If exports start
    returning 403/404 again, refresh it using the same steps.

    Returns True if the cookie was injected, False if not configured.
    """
    sid = os.environ.get("SCREENER_SESSION", "").strip()
    if not sid:
        log.debug(
            "_screener_inject_session: SCREENER_SESSION not set "
            "— Excel export fallback disabled (HTML scraping still works)"
        )
        return False

    # Inject the session cookie so all subsequent requests are authenticated
    session.cookies.set("sessionid", sid, domain="screener.in")
    log.info(
        "_screener_inject_session: sessionid injected (%.8s...) — export fallback enabled",
        sid,
    )
    return True


def _get_screener_session() -> requests.Session:
    """
    Return a warmed requests.Session for screener.in.

    The first call does:
      GET https://screener.in/  → picks up csrftoken cookie

    If SCREENER_SESSION env var is set (the ``sessionid`` cookie copied from a
    browser session), it is injected so that authenticated endpoints like the
    Excel export (/company/{slug}/export/) become accessible.

    Subsequent calls reuse the same session object.
    """
    global _screener_session, _screener_session_warmed, _screener_session_logged_in

    if _screener_session is None:
        from data.proxy_session import apply_proxy_to_session, proxy_configured
        _screener_session = requests.Session()
        _screener_session.headers.update(_screener_headers())
        apply_proxy_to_session(_screener_session)   # routes via SCRAPERAPI/Fixie if configured
        if not proxy_configured():
            log.warning(
                "_get_screener_session: no proxy configured — screener.in may be "
                "blocked from Railway (set SCRAPERAPI_KEY or FIXIE_URL env var)"
            )

    if not _screener_session_warmed:
        try:
            # Warm up: visit homepage to get cookies
            warm_resp = _screener_session.get(
                "https://screener.in/",
                timeout=10,
                allow_redirects=True,
            )
            _screener_session_warmed = warm_resp.status_code == 200
            if _screener_session_warmed:
                log.debug(
                    "_get_screener_session: warmup OK — cookies: %s",
                    list(_screener_session.cookies.keys()),
                )
                # Inject session cookie if SCREENER_SESSION is configured (non-fatal)
                if not _screener_session_logged_in:
                    _screener_session_logged_in = _screener_inject_session(_screener_session)
            else:
                log.warning(
                    "_get_screener_session: warmup returned HTTP %s",
                    warm_resp.status_code,
                )
        except Exception as exc:
            log.warning("_get_screener_session: warmup failed: %s", exc)

    return _screener_session


def reset_screener_session() -> None:
    """Force a fresh session next time (call after 403/429 to recover)."""
    global _screener_session, _screener_session_warmed, _screener_session_logged_in
    _screener_session = None
    _screener_session_warmed = False
    _screener_session_logged_in = False
    log.info("Screener.in session reset — will re-warm on next request")


# ─── yfinance retry utility ──────────────────────────────────────────────────

# Errors that indicate a transient problem worth retrying.
# - ConnectionError / Timeout: network blip or DNS hiccup
# - HTTPError 429: Yahoo rate-limit (back off and retry)
# - HTTPError 5xx: Yahoo server error (transient)
# - JSONDecodeError: occasionally yfinance returns a malformed response
_YF_RETRYABLE: Tuple[Type[Exception], ...] = (
    requests.exceptions.ConnectionError,
    requests.exceptions.Timeout,
    json.JSONDecodeError,
)

# Errors indicating a permanent failure — retrying would not help.
# - HTTP 404: symbol does not exist in Yahoo Finance
# - HTTP 400: bad request (malformed symbol)
# - ValueError / TypeError from downstream yfinance parsing
_YF_PERMANENT: Tuple[Type[Exception], ...] = (
    ValueError,
    TypeError,
)


def yf_fetch_with_retry(
    fn: Callable[..., Any],
    *args: Any,
    max_retries: int = 3,
    base_delay: float = 1.0,
    **kwargs: Any,
) -> Any:
    """
    Execute a callable (typically a yfinance API call) with exponential-backoff
    retry for transient failures.

    Retry policy
    ------------
    - Retries on: ConnectionError, Timeout, HTTP 429, HTTP 5xx, JSONDecodeError
    - Does NOT retry on: HTTP 404 / 400 (permanent), ValueError, TypeError
    - Delays: base_delay * 2^attempt + uniform jitter [0, 0.5s]
      e.g., with base_delay=1.0: ~1.0s, ~2.0s, ~4.0s

    Usage
    -----
        # .history() call with retry
        df = yf_fetch_with_retry(yf.Ticker("TCS.NS").history, period="1y")

        # .info property via lambda
        info = yf_fetch_with_retry(lambda: yf.Ticker("TCS.NS").info)

    Returns the callable's return value on success.
    Raises the last exception if all retries are exhausted.
    Raises immediately (no retry) for permanent errors.
    """
    last_exc: Exception = RuntimeError("yf_fetch_with_retry: no attempts made")

    for attempt in range(max_retries):
        try:
            return fn(*args, **kwargs)

        except requests.exceptions.HTTPError as exc:
            resp = getattr(exc, "response", None)
            status = resp.status_code if resp is not None else 0
            if status in (429,) or (500 <= status < 600):
                last_exc = exc
                delay = base_delay * (2 ** attempt) + random.uniform(0, 0.5)
                log.debug(
                    "yfinance HTTP %s (attempt %d/%d) — retrying in %.1fs",
                    status, attempt + 1, max_retries, delay,
                )
                time.sleep(delay)
            else:
                # 404, 400, etc. — permanent failure, raise immediately
                raise

        except _PERMANENT as exc:
            raise

        except _YF_RETRYABLE as exc:
            last_exc = exc
            delay = base_delay * (2 ** attempt) + random.uniform(0, 0.5)
            log.debug(
                "yfinance transient %s (attempt %d/%d) — retrying in %.1fs",
                type(exc).__name__, attempt + 1, max_retries, delay,
            )
            time.sleep(delay)

        except Exception as exc:
            # Unknown error — treat as transient for one retry only,
            # then let it propagate (avoids swallowing genuine bugs).
            if attempt < 1:
                last_exc = exc
                delay = base_delay + random.uniform(0, 0.5)
                log.debug(
                    "yfinance unknown error %s (attempt %d/%d) — retrying once in %.1fs",
                    type(exc).__name__, attempt + 1, max_retries, delay,
                )
                time.sleep(delay)
            else:
                raise

    raise last_exc


# Private alias for the permanent-error tuple used in the except clause above.
# Defined after the function because it is only used internally.
_PERMANENT = _YF_PERMANENT


# ─── 1. OHLCV ────────────────────────────────────────────────────────────────

def get_ohlcv(symbol: str, period: str = "1y"):
    """
    Fetch OHLCV data for a symbol using yfinance.

    Args:
        symbol: Ticker symbol, e.g. "RELIANCE.NS", "TCS.NS"
        period: yfinance period string — "1d","5d","1mo","3mo","6mo","1y","2y","5y","10y","ytd","max"

    Returns:
        pandas.DataFrame with columns [Open, High, Low, Close, Volume], or None on failure.
    """
    from data.symbol_map import resolve_yf, is_excluded
    if is_excluded(symbol):
        log.debug("get_ohlcv: skipping excluded symbol %s", symbol)
        return None
    resolved = resolve_yf(symbol)
    if resolved is None:
        log.debug("get_ohlcv: %s resolved to None (excluded)", symbol)
        return None
    if resolved != symbol:
        log.debug("get_ohlcv: symbol resolved %s → %s", symbol, resolved)

    try:
        ticker = yf.Ticker(resolved)
        df = yf_fetch_with_retry(ticker.history, period=period)
        if df.empty:
            log.warning("get_ohlcv: no data returned for %s (period=%s)", resolved, period)
            return None
        df.index = df.index.tz_localize(None)
        return df[["Open", "High", "Low", "Close", "Volume"]]
    except Exception as e:
        log.error("get_ohlcv(%s): %s", resolved, e)
        return None


# ─── 2. NSE FII/DII ──────────────────────────────────────────────────────────

# ── FII/DII helpers ───────────────────────────────────────────────────────────

def _parse_fii_crore(text: str) -> Optional[float]:
    """
    Parse an Indian number string that may include commas, rupee signs,
    en-dashes (−) or minus signs, and return a float in crores.
    Returns None if unparseable.
    """
    if not text:
        return None
    cleaned = (
        text.replace(",", "").replace("₹", "").replace("−", "-")
            .replace("\u2212", "-").replace("(", "-").replace(")", "").strip()
    )
    try:
        return float(cleaned)
    except ValueError:
        return None


def _scrape_fii_table(soup: BeautifulSoup) -> Optional[dict]:
    """
    Generic FII/DII table parser.  Scans all <table> tags for one that has
    both FII and DII columns, then extracts the most recent row.
    Works across Goodreturns, BSE website, and other HTML layouts.
    """
    for tbl in soup.find_all("table"):
        ths = [th.get_text(" ", strip=True).lower() for th in tbl.find_all("th")]
        if not ths:
            # Some tables use <td> in the header row
            first_row = tbl.find("tr")
            if first_row:
                ths = [td.get_text(" ", strip=True).lower()
                       for td in first_row.find_all("td")]

        has_fii = any("fii" in h or "fpi" in h for h in ths)
        has_dii = any("dii" in h or "mf" in h for h in ths)
        if not (has_fii and has_dii):
            continue

        # Map column index → role
        date_idx = fii_net_idx = dii_net_idx = None
        fii_buy_idx = fii_sell_idx = dii_buy_idx = dii_sell_idx = None
        for i, h in enumerate(ths):
            if "date" in h:
                date_idx = i
            elif ("fii" in h or "fpi" in h) and "net" in h:
                fii_net_idx = i
            elif ("fii" in h or "fpi" in h) and "buy" in h:
                fii_buy_idx = i
            elif ("fii" in h or "fpi" in h) and "sell" in h:
                fii_sell_idx = i
            elif "dii" in h and "net" in h:
                dii_net_idx = i
            elif "dii" in h and "buy" in h:
                dii_buy_idx = i
            elif "dii" in h and "sell" in h:
                dii_sell_idx = i

        # Need at least buy+sell or net for each category
        can_fii = fii_net_idx is not None or (
            fii_buy_idx is not None and fii_sell_idx is not None
        )
        can_dii = dii_net_idx is not None or (
            dii_buy_idx is not None and dii_sell_idx is not None
        )
        if not (can_fii and can_dii):
            continue

        # Walk data rows (skip header)
        for row in tbl.find_all("tr")[1:]:
            cells = [td.get_text(" ", strip=True) for td in row.find_all("td")]
            if len(cells) < 3:
                continue
            try:
                if fii_net_idx is not None:
                    fii_net = _parse_fii_crore(cells[fii_net_idx])
                else:
                    b = _parse_fii_crore(cells[fii_buy_idx])
                    s = _parse_fii_crore(cells[fii_sell_idx])
                    fii_net = (b - s) if (b is not None and s is not None) else None

                if dii_net_idx is not None:
                    dii_net = _parse_fii_crore(cells[dii_net_idx])
                else:
                    b = _parse_fii_crore(cells[dii_buy_idx])
                    s = _parse_fii_crore(cells[dii_sell_idx])
                    dii_net = (b - s) if (b is not None and s is not None) else None

                if fii_net is None or dii_net is None:
                    continue

                date_str = cells[date_idx].strip() if date_idx is not None else ""
                return {"date": date_str, "fii_net": fii_net, "dii_net": dii_net}
            except (IndexError, TypeError):
                continue

    return None


# ── Per-source attempt functions ──────────────────────────────────────────────

def _try_nse_fii_dii() -> dict | None:
    """
    Source 1: NSE fiidiiTradeReact JSON — direct session + allorigins proxy fallback.

    NSE requires the `nsit` cookie (JS-generated). Without it the API returns
    200 with an empty body. We try:
      a) two-step warm-up session (homepage → data page → API)
      b) allorigins.win CORS proxy as a fallback — used by open-source NSE
         libraries (MrChartist/fii-dii-data) when direct calls fail.
    """
    api_url  = "https://www.nseindia.com/api/fiidiiTradeReact"
    seed_url = "https://www.nseindia.com"
    warm_url = "https://www.nseindia.com/market-data/fii-dii-activity"
    headers  = {
        **_HEADERS,
        "Accept":          "application/json, text/plain, */*",
        # Omit "Accept-Encoding: br" — NSE honours brotli but requests lacks
        # brotlicffi by default → garbled bytes → silent parse failure.
        # Let requests choose gzip/deflate which it can always decompress.
        "Accept-Encoding": "gzip, deflate",
        "Referer":         warm_url,
        "Cache-Control":   "no-cache",
        "Pragma":          "no-cache",
        "Sec-Fetch-Dest":  "empty",
        "Sec-Fetch-Mode":  "cors",
        "Sec-Fetch-Site":  "same-origin",
        "X-Requested-With": "XMLHttpRequest",
    }

    def _parse_nse_json(text: str) -> Optional[dict]:
        if not text.strip():
            return None
        data = json.loads(text)
        if not data:
            return None

        fii_net: Optional[float] = None
        dii_net: Optional[float] = None
        date_str = ""

        # ── New format (May 2026+) ───────────────────────────────────────────
        # NSE changed the API response schema: instead of a single object with
        # fiiNet/diiNet keys, it now returns an array of category objects:
        #   [{"category": "DII",     "date": "12-May-2026",
        #     "buyValue": "20684.82", "sellValue": "12694.5", "netValue": "7990.32"},
        #    {"category": "FII/FPI", "date": "12-May-2026",
        #     "buyValue": "16555.27", "sellValue": "18514.66", "netValue": "-1959.39"}]
        if isinstance(data, list) and data and "category" in data[0]:
            for item in data:
                cat     = (item.get("category") or "").lower()
                net_raw = item.get("netValue")
                date_str = item.get("date", date_str)
                try:
                    net_val: Optional[float] = float(net_raw) if net_raw is not None else None
                except (TypeError, ValueError):
                    net_val = None
                if any(k in cat for k in ("fii", "fpi", "foreign")):
                    fii_net = net_val
                elif "dii" in cat:
                    dii_net = net_val
            # Both-zero guard: a holiday or failed parse returns 0/0 — skip it
            # so the next source (BSE, Trendlyne …) gets a chance.
            if fii_net is None or (fii_net == 0.0 and dii_net == 0.0):
                return None
            return {
                "date":    date_str,
                "fii_net": fii_net,
                "dii_net": dii_net if dii_net is not None else 0.0,
                "source":  "nse",
            }

        # ── Legacy format ────────────────────────────────────────────────────
        # Single object (or array[0]) with fiiNet / diiNet keys.
        latest  = data[0] if isinstance(data, list) else data
        fii_net = float(latest.get("fiiNet") or latest.get("FII_NET") or 0)
        dii_net = float(latest.get("diiNet") or latest.get("DII_NET") or 0)
        # Guard against stale zero-filled responses
        if fii_net == 0.0 and dii_net == 0.0:
            return None
        return {
            "date":    latest.get("date") or latest.get("DATE") or "",
            "fii_net": fii_net,
            "dii_net": dii_net,
            "source":  "nse",
        }

    # Attempt A: direct session with cookie warm-up
    try:
        session = requests.Session()
        session.get(seed_url, headers=headers, timeout=10)
        time.sleep(1)
        session.get(warm_url, headers=headers, timeout=10)
        resp = session.get(api_url, headers=headers, timeout=10)
        resp.raise_for_status()
        result = _parse_nse_json(resp.text)
        if result:
            return result
    except Exception:
        pass

    # Attempt B: allorigins.win proxy (bypasses cookie requirement)
    try:
        from urllib.parse import quote
        proxy_url = f"https://api.allorigins.win/get?url={quote(api_url)}"
        resp = requests.get(proxy_url, headers=_HEADERS, timeout=15)
        resp.raise_for_status()
        contents = resp.json().get("contents", "")
        result = _parse_nse_json(contents)
        if result:
            result["source"] = "nse_proxy"
            return result
    except Exception:
        pass

    log.warning("FII/DII NSE failed: both direct and proxy attempts returned no data")
    return None


def _try_bse_fii_dii() -> dict | None:
    """
    Source 2: BSE Investor Category-wise Turnover page (confirmed working Apr 2026).

    Previous URL (FiidiiActivity.aspx) returned 404 — BSE moved this data to
    the categorywise_turnover page. This page shows FII/FPI and DII net values
    in an HTML table; no auth or JS required.
    """
    urls = [
        "https://www.bseindia.com/markets/equity/EQReports/categorywise_turnover.aspx",
        "https://www.bseindia.com/markets/equity/EQReports/categorywise_turnover.aspx?expandable=3",
    ]
    headers = {
        **_HEADERS,
        "Accept":  "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Referer": "https://www.bseindia.com/",
    }
    for url in urls:
        try:
            session = requests.Session()
            session.get("https://www.bseindia.com", headers=headers, timeout=10)
            resp = session.get(url, headers=headers, timeout=12)
            resp.raise_for_status()
            soup = BeautifulSoup(resp.text, "html.parser")

            # BSE categorywise page has a specific layout: rows labelled
            # "FII / FPI" and "DII" with buy/sell/net columns.
            result = _scrape_bse_categorywise(soup)
            if result:
                result["source"] = "bse"
                return result

            # Fall back to generic scanner
            result = _scrape_fii_table(soup)
            if result:
                result["source"] = "bse"
                return result
        except Exception as exc:
            log.debug("FII/DII BSE (%s) failed: %s", url, exc)

    log.warning("FII/DII BSE failed: all URL variants exhausted")
    return None


def _scrape_bse_categorywise(soup: BeautifulSoup) -> Optional[dict]:
    """
    BSE categorywise_turnover.aspx specific parser.

    The page has rows like:
      | Category        | Buy (₹Cr) | Sell (₹Cr) | Net (₹Cr) |
      | FII / FPI       | 12345.00  | 14000.00   | -1655.00  |
      | DII             |  8000.00  |  9000.00   | -1000.00  |

    We scan all <tr> elements looking for a row whose first cell contains
    "fii" or "fpi", then grab the net column (last numeric cell or labelled).
    """
    fii_net = dii_net = None
    date_str = ""

    # Try to find a date anywhere on the page
    for tag in soup.find_all(["span", "td", "th", "p"], limit=100):
        txt = tag.get_text(strip=True)
        if re.match(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", txt):
            date_str = txt
            break

    for row in soup.find_all("tr"):
        cells = [td.get_text(" ", strip=True) for td in row.find_all(["td", "th"])]
        if len(cells) < 3:
            continue
        label = cells[0].lower()

        if any(k in label for k in ("fii", "fpi", "foreign")):
            # Try last cell as net, else compute buy-sell
            nums = [_parse_fii_crore(c) for c in cells[1:]]
            nums = [n for n in nums if n is not None]
            if len(nums) >= 3:
                fii_net = nums[-1]               # net is last column
            elif len(nums) == 2:
                fii_net = nums[0] - nums[1]      # buy - sell

        elif any(k in label for k in ("dii", "domestic inst", "mutual fund")):
            nums = [_parse_fii_crore(c) for c in cells[1:]]
            nums = [n for n in nums if n is not None]
            if len(nums) >= 3:
                dii_net = nums[-1]
            elif len(nums) == 2:
                dii_net = nums[0] - nums[1]

    if fii_net is not None and dii_net is not None:
        return {"date": date_str, "fii_net": fii_net, "dii_net": dii_net}
    return None


def _try_trendlyne_fii_dii() -> dict | None:
    """
    Source 3: Trendlyne macro FII/DII page (confirmed working Apr 2026).

    Trendlyne publishes daily FII/DII cash-segment data at a stable URL with
    an HTML table. Confirmed values Apr 22 2026: FII -2078.36 Cr, DII -1048.17 Cr.
    Lower anti-bot aggression than NSE/BSE/Moneycontrol.
    """
    url = "https://trendlyne.com/macro-data/fii-dii/latest/cash-pastmonth/"
    headers = {
        **_HEADERS,
        "Accept":  "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Referer": "https://trendlyne.com/",
    }
    try:
        resp = requests.get(url, headers=headers, timeout=12)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        result = _scrape_fii_table(soup)
        if result:
            result["source"] = "trendlyne"
            return result
        return None
    except Exception as exc:
        log.warning("FII/DII Trendlyne failed: %s", exc)
        return None


def _try_moneycontrol_fii_dii() -> dict | None:
    """
    Source 4: Moneycontrol FII/DII page — last resort, sometimes 403.
    """
    for url in (
        "https://www.moneycontrol.com/stocks/marketinfo/fiidii_activity/",
        "https://www.moneycontrol.com/markets/fii-dii-activity/",
    ):
        headers = {
            **_HEADERS,
            "Accept":          "text/html,application/xhtml+xml,*/*;q=0.8",
            "Referer":         "https://www.moneycontrol.com/markets/",
            "Accept-Language": "en-IN,en;q=0.9",
        }
        try:
            resp = requests.get(url, headers=headers, timeout=12)
            if resp.status_code == 403:
                continue
            resp.raise_for_status()
            soup = BeautifulSoup(resp.text, "html.parser")
            result = _scrape_fii_table(soup)
            if result:
                result["source"] = "moneycontrol"
                return result
        except Exception as exc:
            log.debug("FII/DII Moneycontrol (%s) failed: %s", url, exc)
    log.warning("FII/DII Moneycontrol failed: all URL variants blocked")
    return None


def get_nse_fii_dii() -> dict | None:
    """
    Fetch latest daily FII/DII net flow data (₹ Crores).

    Tries four sources in order until one succeeds:
      1. NSE (direct session + allorigins proxy fallback)
      2. BSE categorywise_turnover.aspx  — confirmed working Apr 2026
      3. Trendlyne HTML table            — confirmed working Apr 2026
      4. Moneycontrol                    — last resort, sometimes 403

    All sources report the same SEBI-filed data. The "source" key is
    informational; callers treat values identically regardless of source.

    Returns {"date", "fii_net", "dii_net", "source"} or None.
    """
    for attempt_fn in (
        _try_nse_fii_dii,
        _try_bse_fii_dii,
        _try_trendlyne_fii_dii,
        _try_moneycontrol_fii_dii,
    ):
        result = attempt_fn()
        if result:
            log.info(
                "FII/DII fetched via %s: FII=%.0f Cr  DII=%.0f Cr",
                result["source"], result["fii_net"], result["dii_net"],
            )
            return result

    log.error("get_nse_fii_dii: all sources (NSE, BSE, Trendlyne, Moneycontrol) failed")
    return None


# ─── 3. MCX PRICES ───────────────────────────────────────────────────────────

def get_mcx_prices() -> dict | None:
    """
    Fetch Gold, Crude Oil, and Silver spot/futures prices via yfinance.

    Returns:
        dict with keys: gold ($/oz), crude ($/bbl), silver ($/oz), or None on failure.
    """
    tickers = {"gold": "GC=F", "crude": "CL=F", "silver": "SI=F"}
    result = {}
    try:
        for name, sym in tickers.items():
            try:
                ticker = yf.Ticker(sym)
                hist = yf_fetch_with_retry(ticker.history, period="1d")
                if not hist.empty:
                    result[name] = round(float(hist["Close"].iloc[-1]), 2)
                else:
                    info = yf_fetch_with_retry(lambda t=ticker: t.info)
                    result[name] = round(float(info.get("regularMarketPrice") or info.get("previousClose") or 0), 2)
            except Exception as inner:
                log.warning("get_mcx_prices: failed for %s (%s): %s", name, sym, inner)
                result[name] = None

        if all(v is None for v in result.values()):
            return None
        return result
    except Exception as e:
        log.error("get_mcx_prices: %s", e)
        return None


# ─── 4. RSS HEADLINES ────────────────────────────────────────────────────────

# Static broad-market feeds (keyword-filtered per symbol inside get_rss_headlines)
# Note: Business Standard and Livemint block automated requests; removed.
_RSS_FEEDS = [
    ("ET Markets",    "https://economictimes.indiatimes.com/markets/rssfeeds/1977021501.cms"),
    ("ET Auto",       "https://economictimes.indiatimes.com/industry/auto/rssfeeds/1286551815.cms"),
    ("Moneycontrol",  "https://www.moneycontrol.com/rss/business.xml"),
    ("Hindu BizLine", "https://www.thehindubusinessline.com/markets/feeder/default.rss"),
]

# Google News RSS — dynamic per symbol; always returns symbol-specific headlines
_GOOGLE_NEWS_RSS_TMPL = (
    "https://news.google.com/rss/search"
    "?q={keyword}+NSE+stock+India"
    "&hl=en-IN&gl=IN&ceid=IN:en"
)


def get_rss_headlines(symbol: str) -> list | None:
    """
    Fetch headlines from static market RSS feeds + a dynamic Google News RSS
    feed for the specific symbol.

    Args:
        symbol: NSE/BSE ticker, e.g. "MARUTI.NS" or "TCS"

    Returns:
        List of dicts [{title, source, published, url}], or None if every
        feed request failed (network-level failure).
    """
    from urllib.parse import quote_plus

    keyword = symbol.replace(".NS", "").replace(".BO", "").strip().upper()

    # Build the full feed list: static feeds + 1 dynamic Google News feed
    feeds_to_try = list(_RSS_FEEDS) + [
        ("Google News", _GOOGLE_NEWS_RSS_TMPL.format(keyword=quote_plus(keyword))),
    ]

    results: list[dict] = []
    any_feed_ok = False

    for source_name, feed_url in feeds_to_try:
        try:
            feed = feedparser.parse(feed_url)
            # bozo=True means malformed XML, but entries may still be present
            if feed.bozo and not feed.entries:
                log.warning("get_rss_headlines: feed parse error for %s", source_name)
                continue
            any_feed_ok = True

            for entry in feed.entries:
                title   = entry.get("title", "").strip()
                summary = entry.get("summary", "")

                # For Google News the feed is already symbol-specific;
                # for static feeds we still keyword-filter.
                is_google = source_name == "Google News"
                if not is_google and (
                    keyword.lower() not in title.lower()
                    and keyword.lower() not in summary.lower()
                ):
                    continue

                published = entry.get("published", "")
                try:
                    if entry.get("published_parsed"):
                        published = str(datetime(*entry.published_parsed[:6]).date())
                except Exception:
                    pass

                results.append({
                    "title":     title,
                    "source":    source_name,
                    "published": published,
                    "url":       entry.get("link", ""),
                })
        except Exception as exc:
            log.warning("get_rss_headlines: error parsing %s: %s", source_name, exc)

    if not any_feed_ok:
        return None
    return results


# ─── 5. SCREENER DATA ────────────────────────────────────────────────────────

def get_screener_data(symbol: str) -> dict | None:
    """
    Scrape key fundamentals from screener.in for an NSE-listed company.

    Args:
        symbol: NSE symbol without exchange suffix, e.g. "RELIANCE", "TCS"

    Returns:
        dict with keys:
          pe                  — Price/Earnings ratio
          revenue_growth      — Revenue YoY growth % (TTM or 1-year)
          ebitda_margin       — EBITDA / OPM margin %
          debt_equity         — Debt/Equity ratio
          roce                — Return on Capital Employed %
          promoter_holding    — Promoter shareholding %
          promoter_pledging   — Promoter pledged shares %
          revenue_growth_qoq  — Quarter-on-quarter revenue growth %
          revenue_cagr_3y     — 3-year compounded sales growth %
          revenue_cagr_5y     — 5-year compounded sales growth %
          eps_cagr_3y         — 3-year compounded profit growth %
          eps_cagr_5y         — 5-year compounded profit growth %
          roe                 — Return on equity % (from top-ratios bar)
          interest_coverage   — Interest coverage ratio (from top-ratios bar)
          ocf_margin          — Operating cash flow / revenue % (from cash flow table)
        Values are floats or None if not found. Returns None on request failure.
    """
    from data.symbol_map import resolve_screener, is_excluded
    if is_excluded(symbol):
        log.debug("get_screener_data: skipping excluded symbol %s", symbol)
        return None
    slug = resolve_screener(symbol)
    if slug != symbol.replace(".NS", "").replace(".BO", "").upper():
        log.debug("get_screener_data: slug resolved %s → %s", symbol, slug)

    # Try consolidated first, then standalone; also retry with original
    # base symbol if the slug override still 404s (graceful fallback chain).
    # Consolidated is preferred because screener.in shows consolidated figures
    # by default for most large-cap companies (e.g. Reliance, HDFC Bank).
    # Standalone-only companies (no consolidated subsidiary) return 404 for
    # the consolidated/ URL, so the fallback to standalone handles them.
    base_fallback = symbol.replace(".NS", "").replace(".BO", "").upper()
    candidates = [slug]
    if base_fallback != slug:
        candidates.append(base_fallback)   # fallback to raw NSE symbol

    resp = None
    _last_status: str = "no_attempt"
    session = _get_screener_session()

    for candidate in candidates:
        for variant in ("consolidated/", ""):   # consolidated first
            url = f"https://www.screener.in/company/{candidate}/{variant}"
            try:
                r = session.get(
                    url,
                    headers=_screener_headers({
                        "Referer": "https://screener.in/",
                        "User-Agent": random.choice(_SCREENER_USER_AGENTS),
                    }),
                    timeout=12,
                )
                _last_status = str(r.status_code)
                if r.status_code == 200:
                    resp = r
                    break
                elif r.status_code in (403, 429):
                    # IP blocked or rate-limited — log clearly and reset session
                    log.warning(
                        "get_screener_data(%s): HTTP %s from screener.in (%s) — "
                        "Railway IP may be blocked. Resetting session.",
                        symbol, r.status_code, url,
                    )
                    reset_screener_session()
                    session = _get_screener_session()
                else:
                    log.info(
                        "get_screener_data(%s): HTTP %s from %s",
                        symbol, r.status_code, url,
                    )
            except requests.exceptions.ConnectionError as req_exc:
                _last_status = "connection_error"
                log.warning(
                    "get_screener_data(%s): connection error for %s: %s",
                    symbol, url, req_exc,
                )
            except requests.exceptions.Timeout:
                _last_status = "timeout"
                log.warning("get_screener_data(%s): timeout hitting %s", symbol, url)
            except Exception as req_exc:
                _last_status = "error"
                log.warning("get_screener_data(%s): request error %s: %s", symbol, url, req_exc)
        if resp is not None:
            break

    if resp is None:
        log.warning(
            "get_screener_data(%s): screener.in unavailable (last status: %s) "
            "— trying Trendlyne tier-2 fallback",
            symbol, _last_status,
        )
        # ── Trendlyne tier-2 fallback ─────────────────────────────────────────
        # Only attempted when screener.in is blocked/down (Railway IP block etc.)
        # Returns same schema as get_screener_data(); no extra wiring needed in agents.
        try:
            from data.trendlyne_fetcher import get_trendlyne_fundamentals
            tl_data = get_trendlyne_fundamentals(symbol)
            if tl_data is not None:
                log.info(
                    "get_screener_data(%s): Trendlyne tier-2 returned data "
                    "(pe=%s roce=%s revg=%s)",
                    symbol, tl_data.get("pe"), tl_data.get("roce"),
                    tl_data.get("revenue_growth"),
                )
                return tl_data
        except Exception as _tl_exc:
            log.warning("get_screener_data(%s): Trendlyne fallback error: %s", symbol, _tl_exc)

        log.warning("get_screener_data(%s): Trendlyne also failed — falling back to yfinance", symbol)
        # ── yfinance tier-3 fallback ──────────────────────────────────────────
        # Resolve to yfinance symbol (try .NS first, then .BO)
        from data.symbol_map import YF_SYMBOL_MAP
        clean_sym = symbol.replace(".NS", "").replace(".BO", "").upper()
        yf_sym = YF_SYMBOL_MAP.get(clean_sym) or f"{clean_sym}.NS"
        fb = _get_yfinance_fundamentals(yf_sym, clean_sym)
        if fb is None and not yf_sym.endswith(".BO"):
            # Try BSE suffix as last resort
            fb = _get_yfinance_fundamentals(f"{clean_sym}.BO", clean_sym)
        if fb is not None:
            log.info("get_screener_data(%s): returning yfinance fallback data", symbol)
        return fb

    try:
        resp.raise_for_status()

        soup = BeautifulSoup(resp.text, "html.parser")
        result = {
            "pe": None,
            "revenue_growth": None,
            "ebitda_margin": None,
            "debt_equity": None,
            "roce": None,
            "promoter_holding": None,
            "promoter_pledging": None,   # % of promoter shares pledged
            "revenue_growth_qoq": None,  # most recent quarter-on-quarter %
            # Tier 3 new fields
            "revenue_cagr_3y": None,     # 3-year compounded sales growth %
            "revenue_cagr_5y": None,     # 5-year compounded sales growth %
            "eps_cagr_3y": None,         # 3-year compounded profit growth %
            "eps_cagr_5y": None,         # 5-year compounded profit growth %
            "roe": None,                 # Return on equity %
            "interest_coverage": None,   # Interest coverage ratio
            # Shareholding breakdown (quarterly snapshot from screener)
            "fii_holding_pct": None,     # FII/FPI % ownership (latest quarter)
            "dii_holding_pct": None,     # DII/MF % ownership (latest quarter)
            "ocf_margin": None,          # Operating cash flow / revenue %
        }

        # ── Top ratios bar ──────────────────────────────────────────────────
        for li in soup.select("#top-ratios li"):
            name_el = li.select_one(".name")
            # Use .number (inner span) not .value (outer span which includes ₹/Cr. text)
            val_el  = li.select_one(".number")
            if not name_el or not val_el:
                continue
            name_txt = name_el.get_text(strip=True).lower()
            val_txt  = val_el.get_text(strip=True).replace(",", "").replace("%", "").strip()
            val = _safe_float(val_txt)
            if "stock p/e" in name_txt or "p/e" == name_txt:
                result["pe"] = val
            elif "debt / equity" in name_txt or "debt/equity" in name_txt:
                result["debt_equity"] = val
            elif "roce" in name_txt:
                result["roce"] = val
            elif "roe" in name_txt and "roce" not in name_txt:
                # ROE label contains "roe" but NOT "roce"
                result["roe"] = val
            elif "interest coverage" in name_txt:
                result["interest_coverage"] = val
            elif "market cap" in name_txt and val:
                # Screener shows in Crores; convert to absolute rupees (matches yfinance.marketCap)
                result["market_cap"] = int(val * 1e7)

        # ── Sector from breadcrumb ──────────────────────────────────────────────
        for a in soup.select('a[href*="/market/"]'):
            txt = a.get_text(strip=True)
            if txt and len(txt) < 50:
                result["sector"] = txt
                break  # first link = broadest sector (e.g. "Energy")

        # ── Shareholding table: promoter, pledging, FII/DII ────────────────────
        # Screener's shareholding section has rows:
        #   Promoters | FIIs | DIIs | Government | Public | ...
        # Each row has one value per quarter; we take the most recent (last) cell.
        found_promoter = False
        for row in soup.select("table.data-table tbody tr"):
            cells = row.find_all("td")
            if not cells:
                continue
            row_label = cells[0].get_text(strip=True).lower()

            def _last_val(cells):
                """Return the last non-None float value across all cells."""
                v_all = [
                    _safe_float(td.get_text(strip=True).replace("%", ""))
                    for td in cells[1:]
                ]
                vals = [v for v in v_all if v is not None]
                return vals[-1] if vals else None

            if not found_promoter and "promoter" in row_label and "pledg" not in row_label:
                result["promoter_holding"] = _last_val(cells)
                found_promoter = True
            elif "pledg" in row_label:
                result["promoter_pledging"] = _last_val(cells)
                # keep scanning — FII/DII rows come after pledging
            elif any(k in row_label for k in ("fii", "fpi", "foreign")):
                result["fii_holding_pct"] = _last_val(cells)
            elif any(k in row_label for k in ("dii", "mutual fund", "insurance", "domestic inst")):
                if result["dii_holding_pct"] is None:
                    result["dii_holding_pct"] = _last_val(cells)

        # ── QoQ revenue growth from quarterly sales table ────────────────────
        # Look for the quarterly P&L section (Sales row) and compute QoQ
        for table in soup.select("table.data-table"):
            header_row = table.select_one("thead tr")
            if not header_row:
                continue
            headers = [th.get_text(strip=True).lower() for th in header_row.find_all("th")]
            # We want quarterly tables (months like "Sep 2023", "Dec 2023")
            if not any("sep" in h or "dec" in h or "mar" in h or "jun" in h for h in headers):
                continue
            for row in table.select("tbody tr"):
                cells = row.find_all("td")
                if not cells:
                    continue
                if "sales" in cells[0].get_text(strip=True).lower():
                    qtrs = []
                    for td in cells[1:]:
                        v = _safe_float(td.get_text(strip=True).replace(",", ""))
                        if v is not None:
                            qtrs.append(v)
                    if len(qtrs) >= 2 and qtrs[-2] > 0:
                        result["revenue_growth_qoq"] = round(
                            (qtrs[-1] - qtrs[-2]) / qtrs[-2] * 100, 2
                        )
                    break
            if result["revenue_growth_qoq"] is not None:
                break

        # ── Revenue growth (CAGR) and Profit growth (CAGR) from section cards ──
        # Single pass over all section.card elements; captures 1yr/3yr/5yr for
        # both Sales Growth and Profit Growth sections.
        for section in soup.select("section.card"):
            header = section.find(["h2", "h3"])
            if not header:
                continue
            header_txt = header.get_text(strip=True).lower()

            if "compounded sales growth" in header_txt:
                # Extract all li items to capture 1yr/3yr/5yr
                for li in section.select("li"):
                    try:
                        txt = li.get_text(strip=True)
                        label = txt.split(":")[0].strip().lower()
                        val_str = re.sub(r"[^\d.\-]", "", txt.split(":")[-1])
                        val = _safe_float(val_str)
                        if "ttm" in label or "1 year" in label:
                            result["revenue_growth"] = val
                        elif "3 year" in label:
                            result["revenue_cagr_3y"] = val
                        elif "5 year" in label:
                            result["revenue_cagr_5y"] = val
                        # "10 years" → ignore (too long-term)
                    except Exception:
                        continue

            elif "compounded profit growth" in header_txt or (
                "profit growth" in header_txt and "sales" not in header_txt
            ):
                # Extract eps/profit CAGR: 1yr/3yr/5yr
                for li in section.select("li"):
                    try:
                        txt = li.get_text(strip=True)
                        label = txt.split(":")[0].strip().lower()
                        val_str = re.sub(r"[^\d.\-]", "", txt.split(":")[-1])
                        val = _safe_float(val_str)
                        if "3 year" in label:
                            result["eps_cagr_3y"] = val
                        elif "5 year" in label:
                            result["eps_cagr_5y"] = val
                        # 1yr / ttm → optional; not used downstream yet
                    except Exception:
                        continue

        # ── OCF margin from annual P&L + Cash Flows sections ────────────────────
        # Scan for:
        #   (a) Annual "Profit & Loss" section → latest Sales figure (denominator)
        #   (b) "Cash Flows" section → latest "Cash from Operating Activity" figure
        # Then compute  ocf_margin = OCF / Sales * 100
        # This distinguishes EBITDA-rich but cash-poor companies (accrual distortion).
        _annual_sales: Optional[float] = None
        _ocf_abs: Optional[float] = None

        for section in soup.select("section.card"):
            header = section.find(["h2", "h3"])
            if not header:
                continue
            htxt = header.get_text(strip=True).lower()

            # Annual P&L card → Sales, OPM%, EPS rows
            if "profit & loss" in htxt or "profit and loss" in htxt:
                # screener.in embeds BOTH annual and quarterly tables in the same
                # section.card (one hidden by CSS). We must skip quarterly tables
                # to avoid overwriting annual revenue_growth with a QoQ figure.
                # Annual FY headers are always "Mar YYYY"; quarterly tables
                # contain Jun / Sep / Dec month names in their headers.
                _QUARTERLY_MONTHS = ("jan", "feb", "apr", "may", "jun",
                                     "jul", "aug", "sep", "oct", "nov", "dec")

                def _row_vals(cells):
                    return [v for v in (
                        _safe_float(td.get_text(strip=True).replace(",", "").replace("%", ""))
                        for td in cells[1:]
                    ) if v is not None]

                # Flag: stop after first sales row found (handles both "Sales" and "Sales+")
                _found_sales_row = False

                for table in section.select("table.data-table"):
                    # Detect quarterly table by checking header column names.
                    # Annual FY headers are "Mar YYYY" only; quarterly tables
                    # contain Jun/Sep/Dec/Jan/Feb etc. — skip those.
                    hdr_row = table.select_one("thead tr")
                    if hdr_row:
                        hdr_texts = [th.get_text(strip=True).lower()
                                     for th in hdr_row.find_all("th")]
                        if any(any(m in h for m in _QUARTERLY_MONTHS)
                               for h in hdr_texts):
                            continue  # skip quarterly table

                    for row in table.select("tbody tr"):
                        cells = row.find_all("td")
                        if not cells:
                            continue
                        row_label = cells[0].get_text(strip=True).lower()

                        # Match "Sales" (standalone) or "Sales+" (consolidated).
                        # Screener.in consolidated pages show only "Sales+";
                        # standalone pages show plain "Sales". Accept either.
                        # Use _found_sales_row flag so we don't double-process.
                        if row_label.startswith("sales") and not _found_sales_row:
                            _found_sales_row = True
                            vals = _row_vals(cells)
                            if vals:
                                _annual_sales = vals[-1]
                                # YoY revenue growth
                                if len(vals) >= 2 and vals[-2] and vals[-2] > 0:
                                    result["revenue_growth"] = round(
                                        (vals[-1] - vals[-2]) / vals[-2] * 100, 2
                                    )
                                # 3yr CAGR
                                if len(vals) >= 4 and vals[-4] and vals[-4] > 0:
                                    result["revenue_cagr_3y"] = round(
                                        ((vals[-1] / vals[-4]) ** (1/3) - 1) * 100, 2
                                    )
                                # 5yr CAGR
                                if len(vals) >= 6 and vals[-6] and vals[-6] > 0:
                                    result["revenue_cagr_5y"] = round(
                                        ((vals[-1] / vals[-6]) ** (1/5) - 1) * 100, 2
                                    )
                        elif "opm" in row_label:
                            vals = _row_vals(cells)
                            if vals:
                                result["ebitda_margin"] = vals[-1]
                        elif "eps" in row_label:
                            vals = _row_vals(cells)
                            if len(vals) >= 4 and vals[-4] and vals[-4] > 0:
                                result["eps_cagr_3y"] = round(
                                    ((vals[-1] / vals[-4]) ** (1/3) - 1) * 100, 2
                                )
                            if len(vals) >= 6 and vals[-6] and vals[-6] > 0:
                                result["eps_cagr_5y"] = round(
                                    ((vals[-1] / vals[-6]) ** (1/5) - 1) * 100, 2
                                )
                    if _annual_sales is not None:
                        break

            # Balance Sheet card → Debt/Equity (Borrowings / (Equity + Reserves))
            elif "balance sheet" in htxt:
                _borrowings = _equity_cap = _reserves = None
                for table in section.select("table.data-table"):
                    for row in table.select("tbody tr"):
                        cells = row.find_all("td")
                        if not cells:
                            continue
                        lbl = cells[0].get_text(strip=True).lower()
                        vals = [v for v in (
                            _safe_float(td.get_text(strip=True).replace(",", ""))
                            for td in cells[1:]
                        ) if v is not None]
                        if not vals:
                            continue
                        if "borrowing" in lbl:
                            _borrowings = vals[-1]
                        elif "equity capital" in lbl:
                            _equity_cap = vals[-1]
                        elif lbl.startswith("reserves"):
                            _reserves = vals[-1]
                if _borrowings is not None and _equity_cap is not None and _reserves is not None:
                    equity_total = _equity_cap + _reserves
                    if equity_total > 0:
                        result["debt_equity"] = round(_borrowings / equity_total, 3)

            # Cash Flows card → row containing "operating" (Cash from Operating Activity)
            elif "cash flows" in htxt or "cash flow" in htxt:
                for table in section.select("table.data-table"):
                    for row in table.select("tbody tr"):
                        cells = row.find_all("td")
                        if not cells:
                            continue
                        row_label = cells[0].get_text(strip=True).lower()
                        if "operating" in row_label:
                            vals = [
                                _safe_float(td.get_text(strip=True).replace(",", ""))
                                for td in cells[1:]
                            ]
                            vals = [v for v in vals if v is not None]
                            if vals:
                                _ocf_abs = vals[-1]
                            break
                    if _ocf_abs is not None:
                        break

        if _annual_sales and _annual_sales > 0 and _ocf_abs is not None:
            result["ocf_margin"] = round(_ocf_abs / _annual_sales * 100, 2)

        return result
    except Exception as e:
        log.error("get_screener_data(%s): %s", symbol, e)
        return None


def _safe_float(val: str) -> float | None:
    """Convert a string to float, return None if not possible."""
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return None


# ─── 5b. yfinance fundamentals fallback ──────────────────────────────────────

def _get_yfinance_fundamentals(yf_symbol: str, display_symbol: str) -> dict | None:
    """
    Extract fundamental ratios from yfinance Ticker.info as a fallback when
    screener.in is unavailable (blocked IP, timeout, 403, etc.).

    Maps yfinance fields to the same output schema as get_screener_data() so
    agents work without code changes.  Fields that yfinance doesn't provide
    (CAGR series, promoter pledging, FII/DII split) are returned as None.

    Returns the same schema as get_screener_data() with two additions:
      data_source      — "yfinance_fallback"  (agents can check data quality)
      sector           — sector string from yfinance (used by macro sector adj)

    Returns None if yfinance also fails.
    """
    try:
        ticker = yf.Ticker(yf_symbol)
        info   = ticker.info or {}

        if not info or info.get("quoteType") not in (
            "EQUITY", "MUTUALFUND", None
        ):
            # Empty info dict or a non-equity type we can't use
            log.debug("_get_yfinance_fundamentals(%s): empty or non-equity info", yf_symbol)
            return None

        def _pct(val, mult=100.0):
            """Convert decimal fraction → percentage, return None if unavailable."""
            try:
                f = float(val)
                return round(f * mult, 2) if f != 0 else None
            except (TypeError, ValueError):
                return None

        def _raw(val):
            try:
                f = float(val)
                return round(f, 4) if f != 0 else None
            except (TypeError, ValueError):
                return None

        # ── P/E ───────────────────────────────────────────────────────────────
        pe = _raw(info.get("trailingPE") or info.get("forwardPE"))

        # ── Revenue growth (YoY %) ────────────────────────────────────────────
        # yfinance revenueGrowth is a decimal (0.12 = 12%)
        rev_growth = _pct(info.get("revenueGrowth"))

        # ── EBITDA / OPM margin ───────────────────────────────────────────────
        # operatingMargins is a better proxy for OPM than ebitdaMargins
        ebitda_margin = _pct(info.get("operatingMargins") or info.get("ebitdaMargins"))

        # ── Debt / Equity ─────────────────────────────────────────────────────
        # yfinance debtToEquity is already a ratio (e.g. 0.45 = 45%), divide by 100
        de_raw = info.get("debtToEquity")
        debt_equity = round(float(de_raw) / 100.0, 3) if de_raw not in (None, 0) else None

        # ── ROCE — approximated as EBIT / (Total Assets - Current Liabilities) ─
        # yfinance doesn't give ROCE directly; use returnOnAssets as a rough proxy
        roce = _pct(info.get("returnOnAssets"))

        # ── ROE ───────────────────────────────────────────────────────────────
        roe = _pct(info.get("returnOnEquity"))

        # ── Promoter holding — yfinance heldPercentInsiders (US-centric) ──────
        # For Indian stocks this is often 0 or very low; treat as best-effort
        promoter_holding = _pct(info.get("heldPercentInsiders"))

        # ── Institutional holding — split unavailable; use total institutions ──
        # For Indian stocks: FII+DII combined from heldPercentInstitutions
        inst_total = _pct(info.get("heldPercentInstitutions"))
        # Split 60/40 FII/DII as a rough estimate (India average is ~60% FII of inst.)
        fii_est = round(inst_total * 0.60, 2) if inst_total else None
        dii_est = round(inst_total * 0.40, 2) if inst_total else None

        # ── OCF margin ────────────────────────────────────────────────────────
        ocf_abs  = info.get("operatingCashflow") or info.get("freeCashflow")
        rev_abs  = info.get("totalRevenue")
        ocf_margin = None
        if ocf_abs and rev_abs and float(rev_abs) > 0:
            ocf_margin = round(float(ocf_abs) / float(rev_abs) * 100, 2)

        # ── Earnings growth ───────────────────────────────────────────────────
        # earningsGrowth is YoY (decimal); map to eps_cagr_3y as best-effort
        eps_cagr_approx = _pct(info.get("earningsGrowth"))

        result = {
            "pe":               pe,
            "revenue_growth":   rev_growth,
            "ebitda_margin":    ebitda_margin,
            "debt_equity":      debt_equity,
            "roce":             roce,
            "roe":              roe,
            "promoter_holding": promoter_holding,
            "promoter_pledging": None,          # not available in yfinance
            "revenue_growth_qoq": None,         # not in yfinance.info
            "revenue_cagr_3y":  None,           # not in yfinance.info
            "revenue_cagr_5y":  None,
            "eps_cagr_3y":      eps_cagr_approx,  # YoY as rough proxy
            "eps_cagr_5y":      None,
            "interest_coverage": None,          # not in yfinance.info
            "fii_holding_pct":  fii_est,        # estimated (inst × 0.60)
            "dii_holding_pct":  dii_est,        # estimated (inst × 0.40)
            "ocf_margin":       ocf_margin,
            # Extra context fields
            "sector":           info.get("sector") or info.get("industry"),
            "market_cap":       info.get("marketCap"),
            "data_source":      "yfinance_fallback",
        }

        # Validate: at least pe or roe or revenue_growth must be non-None
        if all(result.get(k) is None for k in ("pe", "roe", "revenue_growth", "ebitda_margin")):
            log.warning(
                "_get_yfinance_fundamentals(%s): all key fields None — info incomplete",
                yf_symbol,
            )
            return None

        log.info(
            "_get_yfinance_fundamentals(%s): pe=%.1f roe=%.1f revg=%.1f%% (fallback)",
            display_symbol,
            pe or 0, roe or 0, rev_growth or 0,
        )
        return result

    except Exception as exc:
        log.warning("_get_yfinance_fundamentals(%s): %s", yf_symbol, exc)
        return None


# ─── 6. INR/USD ──────────────────────────────────────────────────────────────

def get_inr_usd() -> float | None:
    """
    Fetch the current USD → INR exchange rate via yfinance (USDINR=X).

    Returns:
        float (e.g. 83.42), or None on failure.
    """
    try:
        ticker = yf.Ticker("USDINR=X")
        hist = yf_fetch_with_retry(ticker.history, period="1d")
        if not hist.empty:
            return round(float(hist["Close"].iloc[-1]), 4)
        info = yf_fetch_with_retry(lambda: ticker.info)
        rate = info.get("regularMarketPrice") or info.get("previousClose")
        return round(float(rate), 4) if rate else None
    except Exception as e:
        log.error("get_inr_usd: %s", e)
        return None


# ─── 7. INDIA VIX ────────────────────────────────────────────────────────────

def get_india_vix() -> float | None:
    """
    Fetch the India VIX index value via yfinance (^INDIAVIX).

    Returns:
        float (e.g. 13.45), or None on failure.
    """
    try:
        ticker = yf.Ticker("^INDIAVIX")
        hist = yf_fetch_with_retry(ticker.history, period="1d")
        if not hist.empty:
            return round(float(hist["Close"].iloc[-1]), 2)
        info = yf_fetch_with_retry(lambda: ticker.info)
        vix = info.get("regularMarketPrice") or info.get("previousClose")
        return round(float(vix), 2) if vix else None
    except Exception as e:
        log.error("get_india_vix: %s", e)
        return None


# ─── 8. SCREENER HISTORICAL DATA ─────────────────────────────────────────────

def _parse_screener_excel(excel_bytes: bytes, symbol: str) -> dict | None:
    """
    Parse the 'Data Sheet' from a screener.in Excel export into the same dict
    schema as get_screener_history().

    screener.in exports a single workbook with multiple visual sheets (Profit &
    Loss, Balance Sheet, …) that use merged cells — all values read as None in
    openpyxl.  The actual machine-readable data lives in the hidden 'Data Sheet'
    tab, which has a flat row-label layout:

      Row N:   "PROFIT & LOSS"         ← section marker
      Row N+1: "Report Date"           ← datetime objects, one per annual column
      Row N+2: "Sales"                 ← revenue (Cr)
      …        expense rows (raw mat, power, employee, …)
      Row N+?: "Other Income"
      Row N+?: "Depreciation"
      Row N+?: "Interest"
      Row N+?: "Profit before tax"
      Row N+?: "Net profit"
      …
      Row M:   "CASH FLOW:"            ← section marker
      Row M+?: "Cash from Investing Activity"  (negative = outflow)
      …
      Row K:   "DERIVED:"              ← section marker
      Row K+1: "Adjusted Equity Shares in Cr"

    ROCE, ROE, Promoter Holding are NOT present in the export.
    OPM is computed as: (PBT + Interest + Depreciation − Other Income) / Sales × 100

    Returns None on parse failure.
    """
    try:
        import io
        import openpyxl  # type: ignore
    except ImportError:
        log.debug("_parse_screener_excel: openpyxl not installed — skipping Excel fallback")
        return None

    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
    except Exception as exc:
        log.warning("_parse_screener_excel(%s): failed to open workbook: %s", symbol, exc)
        return None

    # ── Locate 'Data Sheet' tab ───────────────────────────────────────────────
    ws = None
    for sheet_name in wb.sheetnames:
        if "data" in sheet_name.lower():
            ws = wb[sheet_name]
            break
    if ws is None:
        log.warning(
            "_parse_screener_excel(%s): 'Data Sheet' tab not found — sheets: %s",
            symbol, wb.sheetnames,
        )
        return None

    # ── Read all rows into a flat list ────────────────────────────────────────
    rows: list[list] = [[cell.value for cell in row] for row in ws.iter_rows()]
    if not rows:
        log.warning("_parse_screener_excel(%s): Data Sheet is empty", symbol)
        return None

    # ── Helper: datetime → "Mar 2017" ─────────────────────────────────────────
    _MONTH_ABBR = {
        1: "Jan", 2: "Feb",  3: "Mar",  4: "Apr",
        5: "May", 6: "Jun",  7: "Jul",  8: "Aug",
        9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
    }

    def _dt_to_year_str(v) -> Optional[str]:
        """Convert a datetime / date / 'Mar 2017' string to a normalised year label."""
        from datetime import date as _date, datetime as _datetime
        if isinstance(v, (_datetime, _date)):
            return f"{_MONTH_ABBR.get(v.month, 'Mar')} {v.year}"
        if isinstance(v, str):
            v = v.strip()
            if re.match(r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4}", v):
                return v
        return None

    def _to_float(v) -> Optional[float]:
        if v is None:
            return None
        try:
            return float(str(v).replace(",", "").replace("%", "").strip())
        except (ValueError, TypeError):
            return None

    # ── Find the P&L 'Report Date' row (annual column headers) ───────────────
    # The P&L section opens with a row whose col-A is 'PROFIT & LOSS' (or similar).
    # The very next 'Report Date' row gives the column → year mapping.
    pnl_report_idx: Optional[int] = None
    in_pnl_section = False
    for i, row in enumerate(rows):
        label = str(row[0] or "").strip().upper() if row else ""
        if "PROFIT" in label and "LOSS" in label:
            in_pnl_section = True
            continue
        if in_pnl_section and "REPORT DATE" in label:
            pnl_report_idx = i
            break

    # Fallback: take the very first 'Report Date' row anywhere
    if pnl_report_idx is None:
        for i, row in enumerate(rows):
            if row and "REPORT DATE" in str(row[0] or "").strip().upper():
                pnl_report_idx = i
                break

    if pnl_report_idx is None:
        log.warning(
            "_parse_screener_excel(%s): no 'Report Date' row found in Data Sheet",
            symbol,
        )
        return None

    # ── Extract year column indices and labels ────────────────────────────────
    report_row = rows[pnl_report_idx]
    years: list[str] = []
    year_col_indices: list[int] = []
    for j, v in enumerate(report_row):
        if j == 0:              # skip label cell
            continue
        yr = _dt_to_year_str(v)
        if yr:
            years.append(yr)
            year_col_indices.append(j)

    if not years:
        log.warning(
            "_parse_screener_excel(%s): Report Date row has no recognisable year values",
            symbol,
        )
        return None

    n_cols = len(years)

    # ── Row extraction helpers ────────────────────────────────────────────────
    def _vals_for_row(row: list) -> list[Optional[float]]:
        """Extract the n_cols values aligned to year_col_indices."""
        return [_to_float(row[j]) if j < len(row) else None for j in year_col_indices]

    def _find_row_exact(label: str, start: int = 0, end: Optional[int] = None) -> Optional[list]:
        """Return values for the first row whose col-A equals label (case-insensitive)."""
        target = label.lower().strip()
        for row in rows[start: end]:
            if row and str(row[0] or "").strip().lower() == target:
                return _vals_for_row(row)
        return None

    def _find_row_contains(substr: str, start: int = 0, end: Optional[int] = None) -> Optional[list]:
        """Return values for the first row whose col-A contains substr."""
        target = substr.lower()
        for row in rows[start: end]:
            if row and target in str(row[0] or "").lower():
                return _vals_for_row(row)
        return None

    # ── Locate section boundaries ─────────────────────────────────────────────
    pnl_start = pnl_report_idx + 1
    # Stop P&L search at the next major section (Quarters or Balance Sheet)
    pnl_end = len(rows)
    for i in range(pnl_start, len(rows)):
        label_up = str(rows[i][0] or "").strip().upper() if rows[i] else ""
        if "QUARTER" in label_up or "BALANCE SHEET" in label_up:
            pnl_end = i
            break

    # Find CASH FLOW section start
    cf_start = 0
    for i, row in enumerate(rows):
        if row and "CASH FLOW" in str(row[0] or "").strip().upper():
            cf_start = i
            break

    # Find DERIVED section start
    derived_start = 0
    for i, row in enumerate(rows):
        if row and "DERIVED" in str(row[0] or "").strip().upper():
            derived_start = i
            break

    # ── Extract P&L rows ──────────────────────────────────────────────────────
    sales         = _find_row_contains("sales",              pnl_start, pnl_end)
    other_income  = _find_row_contains("other income",       pnl_start, pnl_end)
    depreciation  = _find_row_contains("depreciation",       pnl_start, pnl_end)
    interest      = _find_row_contains("interest",           pnl_start, pnl_end)
    pbt           = _find_row_contains("profit before tax",  pnl_start, pnl_end)
    net_profit    = _find_row_contains("net profit",         pnl_start, pnl_end)

    # ── Extract Cash Flow row ─────────────────────────────────────────────────
    cash_investing = _find_row_contains("cash from investing", cf_start)

    # ── Extract Adjusted Equity Shares (for EPS) ──────────────────────────────
    equity_shares = (
        _find_row_contains("adjusted equity shares", derived_start)
        if derived_start
        else _find_row_contains("adjusted equity shares")
    )

    # ── Compute OPM% ─────────────────────────────────────────────────────────
    # OPM = (PBT + Interest + Depreciation − Other Income) / Sales × 100
    # This equals EBITDA margin adjusted for other income, matching screener.in's display.
    ebitda_margins: list[Optional[float]] = []
    if sales and pbt and interest and depreciation:
        for idx in range(n_cols):
            s   = sales[idx]
            p   = pbt[idx]
            i_v = interest[idx]
            d   = depreciation[idx]
            oi  = (other_income[idx] if other_income else None) or 0.0
            if s and p is not None and i_v is not None and d is not None and s != 0:
                ebitda_margins.append(round((p + i_v + d - oi) / s * 100, 2))
            else:
                ebitda_margins.append(None)

    # ── Compute EPS ───────────────────────────────────────────────────────────
    # EPS (Rs) = Net Profit (Cr) / Adjusted Equity Shares (Cr)
    # Both are in crores of rupees / crores of shares → result is Rs per share.
    eps_history: list[Optional[float]] = []
    if net_profit and equity_shares:
        for idx in range(n_cols):
            np_v = net_profit[idx]
            es_v = equity_shares[idx]
            if np_v is not None and es_v and es_v != 0:
                eps_history.append(round(np_v / es_v, 2))
            else:
                eps_history.append(None)

    # ── Capex = |Cash from Investing Activity| ───────────────────────────────
    capex_history: list[Optional[float]] = []
    if cash_investing:
        capex_history = [abs(v) if v is not None else None for v in cash_investing]

    log.info(
        "_parse_screener_excel(%s): parsed %d years from Data Sheet",
        symbol, n_cols,
    )
    return {
        "years":                     years,
        "revenue_history":           sales or [],
        "ebitda_margins":            ebitda_margins,
        "pat_history":               net_profit or [],
        "eps_history":               eps_history,
        "depreciation_history":      depreciation or [],
        "capex_history":             capex_history,
        "roce_history":              [],   # not present in screener.in export
        "roe_history":               [],   # not present in screener.in export
        "dividend_payout_history":   [],   # not present in screener.in export
        "promoter_holding_history":  [],   # not present in screener.in export
        "promoter_holding_quarters": [],   # not present in screener.in export
        "years_available":           n_cols,
    }


def get_screener_history(symbol: str) -> dict | None:
    """
    Fetch multi-year historical financial data from screener.in for a company.

    Parses four sections from the company HTML page:
      - Profit & Loss: revenue, OPM%, net profit, EPS, depreciation series
      - Cash Flows:    investing cash flow (capex proxy, absolute value)
      - Ratios:        ROCE%, ROE, dividend payout series
      - Shareholding:  promoter holding quarterly trend

    Args:
        symbol: NSE symbol without exchange suffix, e.g. "RELIANCE", "TCS"

    Returns:
        dict with keys:
          years                    — list of year strings, oldest first, e.g. ["Mar 2015", ...]
          revenue_history          — annual revenue in Cr (list[float|None])
          ebitda_margins           — OPM % series (list[float|None])
          pat_history              — Net Profit in Cr (list[float|None])
          eps_history              — EPS series (list[float|None])
          depreciation_history     — Depreciation in Cr (list[float|None])
          capex_history            — abs(investing CF) in Cr (list[float|None])
          roce_history             — ROCE % annual series (list[float|None])
          roe_history              — ROE % annual series (list[float|None])
          dividend_payout_history  — Dividend payout % series (list[float|None])
          promoter_holding_history — Promoter % quarterly trend (list[float|None])
          promoter_holding_quarters — quarter labels for promoter series (list[str])
          years_available          — int, length of years list
        Returns None on request failure. Never raises.
    """
    from data.symbol_map import resolve_screener, is_excluded

    if is_excluded(symbol):
        log.debug("get_screener_history: skipping excluded symbol %s", symbol)
        return None

    slug = resolve_screener(symbol)
    base_fallback = symbol.replace(".NS", "").replace(".BO", "").upper()
    candidates = [slug]
    if base_fallback != slug:
        candidates.append(base_fallback)

    resp = None
    _last_status: str = "no_attempt"
    session = _get_screener_session()

    for candidate in candidates:
        for variant in ("consolidated/", ""):   # consolidated first
            url = f"https://www.screener.in/company/{candidate}/{variant}"
            try:
                r = session.get(
                    url,
                    headers=_screener_headers({
                        "Referer": "https://screener.in/",
                        "User-Agent": random.choice(_SCREENER_USER_AGENTS),
                    }),
                    timeout=15,
                )
                _last_status = str(r.status_code)
                if r.status_code == 200:
                    resp = r
                    break
                elif r.status_code in (403, 429):
                    log.warning(
                        "get_screener_history(%s): HTTP %s — Railway IP may be blocked",
                        symbol, r.status_code,
                    )
                    reset_screener_session()
                    session = _get_screener_session()
                else:
                    log.info(
                        "get_screener_history(%s): HTTP %s from %s",
                        symbol, r.status_code, url,
                    )
            except Exception as req_exc:
                _last_status = type(req_exc).__name__
                log.warning("get_screener_history: request error %s: %s", url, req_exc)
        if resp is not None:
            break

    if resp is None:
        # ── Excel export fallback (DB-10) ─────────────────────────────────────
        # The export URL is https://www.screener.in/user/company/export/{export_id}/
        # where export_id is embedded in the HTML page as a formaction attribute.
        # We cannot retrieve the export_id without a successful HTML page fetch, so
        # this fallback can only fire when we have the page HTML.  When HTML is
        # blocked entirely (403/429) we have no path forward.
        if not _screener_session_logged_in:
            log.info(
                "get_screener_history(%s): HTML variants failed (last status: %s). "
                "Excel export also unavailable (set SCREENER_SESSION env var to enable; "
                "get value from DevTools → Application → Cookies → screener.in → sessionid).",
                symbol, _last_status,
            )
        else:
            log.warning(
                "get_screener_history(%s): HTML variants failed (last status: %s). "
                "Excel export requires the HTML page to obtain export_id — "
                "cannot fall back to Excel when HTML is blocked.",
                symbol, _last_status,
            )
        return None

    try:
        soup = BeautifulSoup(resp.text, "html.parser")

        def _clean_cell(td) -> Optional[float]:
            """Extract plain text from a cell and parse as float."""
            raw = td.get_text(strip=True)
            cleaned = raw.replace(",", "").replace("%", "").strip()
            return _safe_float(cleaned)

        def _extract_table_rows(section) -> dict[str, list]:
            """
            Parse all data rows in a section's data-table.
            Returns {row_label_lower: [cell_values_as_float_or_None], ...}
            Also returns the header years under key '__headers__'.
            Skips the TTM column if the last header is 'TTM'.
            """
            result: dict[str, list] = {}
            for table in section.select("table.data-table"):
                # Parse headers
                header_row = table.select_one("thead tr")
                if not header_row:
                    continue
                headers = [th.get_text(strip=True) for th in header_row.find_all("th")]
                if not headers:
                    continue

                # Determine if last column is TTM — skip it if so
                skip_last = headers[-1].strip().upper() == "TTM"
                # First header is usually the row label column; data starts at index 1
                data_headers = headers[1:]
                if skip_last and data_headers:
                    data_headers = data_headers[:-1]

                result["__headers__"] = data_headers

                for row in table.select("tbody tr"):
                    cells = row.find_all("td")
                    if not cells:
                        continue
                    label = cells[0].get_text(strip=True).lower().strip()
                    # Data cells (skip label cell, skip TTM if applicable)
                    data_cells = cells[1:]
                    if skip_last and data_cells:
                        data_cells = data_cells[:-1]
                    values = [_clean_cell(td) for td in data_cells]
                    result[label] = values

                # Use first table found in the section
                break
            return result

        def _find_row(rows: dict[str, list], *keywords: str) -> Optional[list]:
            """
            Find a row whose label contains ALL of the given keywords (case-insensitive).
            Returns the value list or None if not found.
            """
            for label, values in rows.items():
                if label == "__headers__":
                    continue
                label_lower = label.lower()
                if all(kw.lower() in label_lower for kw in keywords):
                    return values
            return None

        def _find_row_any(rows: dict[str, list], *keyword_groups) -> Optional[list]:
            """
            Find a row whose label contains ANY of the keyword groups (each group is a tuple).
            Returns the first match's value list.
            """
            for group in keyword_groups:
                if isinstance(group, str):
                    group = (group,)
                result = _find_row(rows, *group)
                if result is not None:
                    return result
            return None

        # Initialise output containers
        years: list[str] = []
        revenue_history: list[Optional[float]] = []
        ebitda_margins: list[Optional[float]] = []
        pat_history: list[Optional[float]] = []
        eps_history: list[Optional[float]] = []
        depreciation_history: list[Optional[float]] = []
        capex_history: list[Optional[float]] = []
        roce_history: list[Optional[float]] = []
        roe_history: list[Optional[float]] = []
        dividend_payout_history: list[Optional[float]] = []
        promoter_holding_history: list[Optional[float]] = []
        promoter_holding_quarters: list[str] = []

        # ── Iterate over all section.card elements ────────────────────────────
        for section in soup.select("section.card"):
            header_el = section.find(["h2", "h3"])
            if not header_el:
                continue
            header_txt = header_el.get_text(strip=True).lower()

            # ── Profit & Loss section ─────────────────────────────────────────
            if "profit" in header_txt and ("loss" in header_txt or "&" in header_txt or "and" in header_txt):
                rows = _extract_table_rows(section)
                hdrs = rows.get("__headers__", [])
                if hdrs and not years:
                    years = hdrs

                # Sales / Revenue
                rev_row = _find_row_any(rows, ("sales",), ("revenue",))
                if rev_row is not None:
                    revenue_history = rev_row

                # OPM % (EBITDA margin)
                opm_row = _find_row_any(rows, ("opm %",), ("opm%",), ("operating profit margin",))
                if opm_row is not None:
                    ebitda_margins = opm_row

                # Net Profit / PAT
                pat_row = _find_row_any(rows, ("net profit",), ("profit after tax",), ("pat",))
                if pat_row is not None:
                    pat_history = pat_row

                # EPS
                eps_row = _find_row_any(rows, ("eps",), ("earnings per share",))
                if eps_row is not None:
                    eps_history = eps_row

                # Depreciation
                dep_row = _find_row_any(rows, ("depreciation",), ("dep.",), ("depr.",))
                if dep_row is not None:
                    depreciation_history = dep_row

            # ── Cash Flow section ─────────────────────────────────────────────
            elif "cash flow" in header_txt or "cash flows" in header_txt:
                rows = _extract_table_rows(section)
                if not years:
                    hdrs = rows.get("__headers__", [])
                    if hdrs:
                        years = hdrs

                # Investing CF (capex proxy) — take absolute value
                inv_row = _find_row_any(
                    rows,
                    ("cash from investing",),
                    ("investing activities",),
                    ("investing",),
                )
                if inv_row is not None:
                    capex_history = [abs(v) if v is not None else None for v in inv_row]

            # ── Ratios / Key Metrics section ──────────────────────────────────
            elif "ratio" in header_txt or "key metric" in header_txt:
                rows = _extract_table_rows(section)
                if not years:
                    hdrs = rows.get("__headers__", [])
                    if hdrs:
                        years = hdrs

                # ROCE %
                roce_row = _find_row_any(rows, ("roce %",), ("roce",))
                if roce_row is not None:
                    roce_history = roce_row

                # ROE
                roe_row = _find_row(rows, "roe")
                if roe_row is not None:
                    roe_history = roe_row

                # Dividend payout
                div_row = _find_row_any(
                    rows,
                    ("dividend payout",),
                    ("dividend payout %",),
                    ("payout %",),
                )
                if div_row is not None:
                    dividend_payout_history = div_row

            # ── Shareholding section ──────────────────────────────────────────
            elif "shareholding" in header_txt or "share holding" in header_txt:
                for table in section.select("table.data-table"):
                    # Parse header quarters
                    header_row = table.select_one("thead tr")
                    if not header_row:
                        continue
                    all_ths = [th.get_text(strip=True) for th in header_row.find_all("th")]
                    # First col is label; rest are quarters
                    quarters = all_ths[1:] if len(all_ths) > 1 else []

                    for row in table.select("tbody tr"):
                        cells = row.find_all("td")
                        if not cells:
                            continue
                        label = cells[0].get_text(strip=True).lower()
                        if "promoter" in label and "pledg" not in label:
                            data_cells = cells[1:]
                            values = [_clean_cell(td) for td in data_cells]
                            promoter_holding_history = values
                            promoter_holding_quarters = quarters
                            break
                    break  # use first table in shareholding section

        # ── Align all annual series to the years list ─────────────────────────
        # If some series are shorter than years (e.g. ratios section has fewer
        # columns), pad from the left with None so indices align with years.
        def _align_to_years(series: list, target_len: int) -> list:
            if not series:
                return [None] * target_len
            if len(series) >= target_len:
                return series[-target_len:]
            return [None] * (target_len - len(series)) + series

        n = len(years)
        if n > 0:
            revenue_history         = _align_to_years(revenue_history, n)
            ebitda_margins          = _align_to_years(ebitda_margins, n)
            pat_history             = _align_to_years(pat_history, n)
            eps_history             = _align_to_years(eps_history, n)
            depreciation_history    = _align_to_years(depreciation_history, n)
            capex_history           = _align_to_years(capex_history, n)
            roce_history            = _align_to_years(roce_history, n)
            roe_history             = _align_to_years(roe_history, n)
            dividend_payout_history = _align_to_years(dividend_payout_history, n)

        html_result = {
            "years":                    years,
            "revenue_history":          revenue_history,
            "ebitda_margins":           ebitda_margins,
            "pat_history":              pat_history,
            "eps_history":              eps_history,
            "depreciation_history":     depreciation_history,
            "capex_history":            capex_history,
            "roce_history":             roce_history,
            "roe_history":              roe_history,
            "dividend_payout_history":  dividend_payout_history,
            "promoter_holding_history": promoter_holding_history,
            "promoter_holding_quarters": promoter_holding_quarters,
            "years_available":          n,
        }

        # ── Excel export supplement (DB-10) ──────────────────────────────────
        # If HTML parsing gave sparse data (< 5 years) and a session cookie is
        # configured, try the authenticated Excel export.  The export URL is
        # embedded in the page HTML as: formaction="/user/company/export/{id}/"
        # We use the csrftoken cookie as the X-CSRFToken header (Django CSRF for
        # non-browser clients).
        if _screener_session_logged_in and n < 5:
            export_id_m = re.search(
                r'formaction=["\']?/user/company/export/(\d+)/', resp.text
            )
            if export_id_m:
                export_id  = export_id_m.group(1)
                export_url = f"https://www.screener.in/user/company/export/{export_id}/"
                csrf_token = session.cookies.get("csrftoken", domain=".screener.in") \
                          or session.cookies.get("csrftoken")
                try:
                    xr = session.post(
                        export_url,
                        headers=_screener_headers({
                            "Referer":      resp.url,
                            "X-CSRFToken":  csrf_token or "",
                            "Accept":       (
                                "application/vnd.openxmlformats-officedocument"
                                ".spreadsheetml.sheet,*/*"
                            ),
                        }),
                        data={"csrfmiddlewaretoken": csrf_token or ""},
                        timeout=30,
                    )
                    if xr.status_code == 200 and len(xr.content) > 1_000:
                        log.info(
                            "get_screener_history(%s): Excel export returned %d bytes "
                            "(HTML gave only %d years)",
                            symbol, len(xr.content), n,
                        )
                        excel_result = _parse_screener_excel(xr.content, symbol)
                        if excel_result and excel_result.get("years_available", 0) > n:
                            return excel_result
                    else:
                        log.debug(
                            "get_screener_history(%s): Excel export HTTP %s from %s",
                            symbol, xr.status_code, export_url,
                        )
                except Exception as xexc:
                    log.debug(
                        "get_screener_history(%s): Excel export error: %s", symbol, xexc
                    )
            else:
                log.debug(
                    "get_screener_history(%s): export_id not found in page HTML "
                    "(sparse data with %d years — screener may require newer login)",
                    symbol, n,
                )

        return html_result

    except Exception as exc:
        log.error("get_screener_history(%s): %s", symbol, exc)
        return None


# ─── SMOKE TEST ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import json

    print("── INR/USD ──────────────────────")
    print(get_inr_usd())

    print("── India VIX ────────────────────")
    print(get_india_vix())

    print("── MCX Prices ───────────────────")
    print(json.dumps(get_mcx_prices(), indent=2))

    print("── OHLCV: TCS.NS (3mo) ──────────")
    df = get_ohlcv("TCS.NS", period="3mo")
    print(df.tail(3) if df is not None else None)

    print("── NSE FII/DII ──────────────────")
    print(json.dumps(get_nse_fii_dii(), indent=2))

    print("── RSS: TCS ─────────────────────")
    headlines = get_rss_headlines("TCS")
    print(json.dumps(headlines[:3] if headlines else headlines, indent=2))

    print("── Screener: TCS ────────────────")
    print(json.dumps(get_screener_data("TCS"), indent=2))
